#!/usr/bin/env python
"""
Import Attica region social welfare structures from Excel + EKKA data.

Sources:
  - mapAttica-Αποτελέσματα Αναζήτησης.xlsx (1,671 structures from Περιφέρεια Αττικής)
  - Πίνακας.csv (190 EKKA-certified NGO operators with AFM)

Usage:
    python scripts/import_attica_structures.py                    # Full import
    python scripts/import_attica_structures.py --dry-run          # Preview only
    python scripts/import_attica_structures.py --clear-existing   # Clear Attica imports first
"""
import argparse
import csv
import os
import re
import sys
from collections import Counter

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook

from my_project import create_app
from my_project.extensions import db
from my_project.registry.models import Structure, StructureType


# --- Constants ---

TYPE_MAP = {
    'Μονάδες Φροντίδας Προσχολικής Αγωγής και Διαπαιδαγώγησης': 'MFPAD',
    'Μονάδα Φροντίδας Ηλικιωμένων': 'MFH',
    'Κέντρο Δημιουργικής Απασχόλησης Παιδιών': 'KDAP',
    'Στέγη Υποστηριζόμενης Διαβίωσης': 'SYD',
    'Κέντρα Αποθεραπείας-Αποκατάστασης (Κ.Α.Α.-ΚΔΗΦ)': 'KDHF-KAA',
    'Δημοτικός Σταθμός': 'DS',
    'Κέντρο Δημιουργικής Απασχόλησης Παιδιών και Ατόμων με Αναπηρία': 'KDAP-AMEA',
    'Κέντρο Ημερήσιας Φροντίδας Ηλικιωμένων': 'KHFH',
    'Πιστοποιημένος φορέας ΜΚΟ': 'MKO',
}

NEW_TYPES = {
    'DS': ('DS', 'Δημοτικός Παιδικός Σταθμός', 'Δημοτικοί βρεφονηπιακοί και παιδικοί σταθμοί'),
    'KDAP-AMEA': ('KDAP-AMEA', 'ΚΔΑΠ Ατόμων με Αναπηρία', 'Κέντρα Δημιουργικής Απασχόλησης Παιδιών και Ατόμων με Αναπηρία'),
    'KHFH': ('KHFH', 'Κέντρο Ημερήσιας Φροντίδας Ηλικιωμένων', 'Κέντρα ημερήσιας φροντίδας και απασχόλησης ηλικιωμένων'),
    'MKO': ('MKO', 'Πιστοποιημένος Φορέας ΜΚΟ', 'Πιστοποιημένοι φορείς παροχής κοινωνικής φροντίδας (Μη Κερδοσκοπικοί)'),
}

PE_MAP = {
    'ΚΕΝΤΡΙΚΟΣ ΤΟΜΕΑΣ': 'KT',
    'ΒΟΡΕΙΟΣ ΤΟΜΕΑΣ': 'BT',
    'ΝΟΤΙΟΣ ΤΟΜΕΑΣ': 'NT',
    'ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ': 'AA',
    'ΠΕΙΡΑΙΩΣ ΚΑΙ ΝΗΣΩΝ': 'PN',
    'ΔΥΤΙΚΟΣ ΤΟΜΕΑΣ': 'DT',
    'ΔΥΤΙΚΗ ΑΤΤΙΚΗ': 'DA',
}


# --- Helpers ---

def _normalize(name):
    """Normalize Greek org name for fuzzy matching."""
    if not name:
        return ''
    name = name.upper().strip()
    name = re.sub(r'[«»""\'\"]+', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name


def parse_address(raw_address):
    """Parse combined Greek address into (street, postal_code) tuple.

    Examples:
        'ΜΕΣΟΓΕΙΩΝ 500 & ΕΥΚΑΛΥΠΤΩΝ 2, 15342,ΑΓ. ΠΑΡΑΣΚΕΥΗ' → ('ΜΕΣΟΓΕΙΩΝ 500 & ΕΥΚΑΛΥΠΤΩΝ 2', '15342')
        'ΣΟΥΛΙΟΥ 42Β, 15343, ΑΓ. ΠΑΡΑΣΚΕΥΗ' → ('ΣΟΥΛΙΟΥ 42Β', '15343')
        'Λ. ΜΕΣΟΓΕΙΩΝ 583, 153 43 ΣΤΑΥΡΟΣ' → ('Λ. ΜΕΣΟΓΕΙΩΝ 583', '15343')
    """
    if not raw_address:
        return None, None

    postal_match = re.search(r'\b(\d{3})\s?(\d{2})\b', raw_address)
    postal_code = None
    if postal_match:
        postal_code = postal_match.group(1) + postal_match.group(2)

    parts = [p.strip() for p in raw_address.split(',')]
    street = parts[0] if parts else raw_address

    return street, postal_code


def infer_ownership(type_name, structure_name):
    """Infer ownership type from structure type and name."""
    if type_name == 'Δημοτικός Σταθμός':
        return 'municipal'
    if type_name == 'Πιστοποιημένος φορέας ΜΚΟ':
        return 'ngo'

    name_upper = (structure_name or '').upper()
    if 'ΔΗΜΟΣ' in name_upper or 'ΔΗΜΟΤΙΚ' in name_upper or 'ΔΗΜΟΥ' in name_upper:
        return 'municipal'

    return 'private'


# --- Core Functions ---

def ensure_types(dry_run=False):
    """Ensure all required structure types exist, create missing ones."""
    existing = {st.code: st for st in StructureType.query.all()}
    created = []

    for code, (_, name, desc) in NEW_TYPES.items():
        if code not in existing:
            if dry_run:
                print(f"  [DRY] Would create type: {code} — {name}")
                # Create a transient object for dry-run type lookups
                fake_st = StructureType(code=code, name=name, description=desc)
                fake_st.id = -1  # placeholder ID
                existing[code] = fake_st
            else:
                st = StructureType(code=code, name=name, description=desc)
                db.session.add(st)
                created.append(code)

    if created and not dry_run:
        db.session.flush()
        existing = {st.code: st for st in StructureType.query.all()}
        print(f"  Created {len(created)} new structure types: {', '.join(created)}")
    elif not created and not dry_run:
        print(f"  All structure types already exist ({len(existing)} total)")

    return existing


def load_ekka_data(csv_path):
    """Load EKKA certified orgs CSV -> dict of normalized_name -> {afm, legal_form, cert_expiry}."""
    ekka = {}
    if not os.path.exists(csv_path):
        print(f"  Warning: EKKA CSV not found at {csv_path}, skipping enrichment")
        return ekka

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            afm = row.get('ΑΦΜ', '').strip()
            name = row.get('Επωνυμία', '').strip()
            title = row.get('Διακριτικός Τίτλος', '').strip()
            legal_form = row.get('Νομική Μορφή', '').strip()
            cert_expiry = row.get('Διάρκεια Πιστοποίησης Έως', '').strip()

            if not afm or not name:
                continue

            entry = {'afm': afm, 'legal_form': legal_form, 'cert_expiry': cert_expiry}
            ekka[_normalize(name)] = entry
            if title and title != '-':
                ekka[_normalize(title)] = entry

    print(f"  Loaded {len(ekka)} EKKA lookup entries from {os.path.basename(csv_path)}")
    return ekka


def find_ekka_match(foreas_name, ekka):
    """Try to match an Excel operator name to EKKA data."""
    if not foreas_name:
        return None

    normalized = _normalize(foreas_name)

    # Exact match
    if normalized in ekka:
        return ekka[normalized]

    # Substring match (EKKA name contains Excel name or vice versa)
    for ekka_name, data in ekka.items():
        if len(normalized) > 5 and (normalized in ekka_name or ekka_name in normalized):
            return data

    return None


def import_structures(excel_path, types, ekka, dry_run=False, clear_existing=False):
    """Parse Excel and import structures into DB."""
    if not os.path.exists(excel_path):
        print(f"  ERROR: Excel file not found: {excel_path}")
        return

    # Clear existing Attica imports if requested
    if clear_existing and not dry_run:
        pe_codes = list(PE_MAP.values())
        deleted = 0
        for structure in Structure.query.all():
            parts = structure.code.split('-')
            if len(parts) >= 3 and parts[-2] in pe_codes:
                db.session.delete(structure)
                deleted += 1
        if deleted:
            db.session.flush()
            print(f"  Cleared {deleted} previously imported Attica structures")

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    # Track sequence numbers per type-PE combo
    seq_counters = Counter()

    # Pre-load existing codes to check for duplicates
    existing_codes = set(row[0] for row in db.session.query(Structure.code).all())

    stats = {'imported': 0, 'skipped': 0, 'errors': 0, 'ekka_matched': 0}
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):  # skip title + header
        row_num = i + 3

        try:
            # Unpack: #, type, name, foreas, character, capacity, license, address, email, site, phone, dimos, enotita
            if len(row) < 13:
                stats['errors'] += 1
                continue

            _, type_name, name, foreas, character, capacity, license_raw, address, email, site, phone, dimos, enotita = row[:13]

            if not name or not type_name:
                stats['errors'] += 1
                continue

            # Map type
            type_code = TYPE_MAP.get(type_name)
            if not type_code:
                print(f"  Row {row_num}: Unknown type '{type_name}', skipping")
                stats['errors'] += 1
                continue

            if type_code not in types:
                print(f"  Row {row_num}: Type code '{type_code}' not in DB, skipping")
                stats['errors'] += 1
                continue

            # Generate code
            pe_abbrev = PE_MAP.get(enotita, 'XX')
            seq_key = f"{type_code}-{pe_abbrev}"
            seq_counters[seq_key] += 1
            code = f"{type_code}-{pe_abbrev}-{seq_counters[seq_key]:04d}"

            # Check uniqueness
            if code in existing_codes:
                stats['skipped'] += 1
                continue

            # Parse address
            street, postal_code = parse_address(address)

            # Phone -> string, take first number only (field is varchar(20))
            phone_str = None
            if phone:
                raw_phone = str(int(phone)) if isinstance(phone, (int, float)) else str(phone)
                phone_str = raw_phone.split()[0][:20] if raw_phone.strip() else None

            # License — skip if ';;' or empty, truncate to 100 chars (field limit)
            license_number = None
            if license_raw and str(license_raw).strip() not in ('', ';;'):
                license_number = str(license_raw).strip()[:100]

            # Capacity — extract leading number (handles "37 ΚΛΙΝΕΣ" etc.)
            cap = None
            if capacity:
                cap_match = re.match(r'(\d+)', str(capacity).strip())
                if cap_match:
                    cap = int(cap_match.group(1))

            # Ownership
            ownership = infer_ownership(type_name, name)
            if character and str(character).strip():
                char_str = str(character).strip().lower()
                if 'δημόσι' in char_str or 'δημοτικ' in char_str:
                    ownership = 'municipal'
                elif 'ιδιωτικ' in char_str:
                    ownership = 'private'

            # Notes — include website if present
            notes_parts = []
            if site:
                notes_parts.append(f"Website: {site}")
            notes = '\n'.join(notes_parts) if notes_parts else None

            # Representative
            rep_name = str(foreas).strip() if foreas else None
            rep_email = str(email).strip() if email else None
            rep_afm = None

            # EKKA enrichment for MKO entries (match on structure name, not foreas)
            if type_code == 'MKO':
                ekka_match = find_ekka_match(name, ekka)
                if ekka_match:
                    rep_afm = ekka_match['afm']
                    cert_info = f"Πιστοποίηση ΕΚΚΑ έως: {ekka_match['cert_expiry']}" if ekka_match.get('cert_expiry') else None
                    if cert_info:
                        notes_parts.append(cert_info)
                        notes = '\n'.join(notes_parts)
                    stats['ekka_matched'] += 1

            if dry_run:
                if stats['imported'] < 10:
                    print(f"  [DRY] {code} | {name[:50]} | {dimos} | {pe_abbrev}" +
                          (f" | AFM:{rep_afm}" if rep_afm else ""))
                elif stats['imported'] == 10:
                    print(f"  ... (showing first 10 of many)")
                stats['imported'] += 1
                existing_codes.add(code)
                continue

            structure = Structure(
                code=code,
                type_id=types[type_code].id,
                name=str(name).strip(),
                street=street,
                city=str(dimos).strip() if dimos else None,
                postal_code=postal_code,
                representative_name=rep_name,
                representative_afm=rep_afm,
                representative_phone=phone_str,
                representative_email=rep_email,
                capacity=cap,
                status='active',
                ownership=ownership,
                license_number=license_number,
                peripheral_unit=str(enotita).strip() if enotita else None,
                notes=notes,
            )
            db.session.add(structure)
            existing_codes.add(code)
            batch.append(structure)
            stats['imported'] += 1

            # Batch flush every 200 rows
            if len(batch) >= 200:
                try:
                    db.session.flush()
                except Exception as e:
                    db.session.rollback()
                    print(f"  Batch flush error at row {row_num}: {e}")
                    stats['errors'] += len(batch)
                    stats['imported'] -= len(batch)
                batch.clear()

        except Exception as e:
            print(f"  Row {row_num}: ERROR — {e}")
            stats['errors'] += 1

    # Final flush + commit
    if batch and not dry_run:
        try:
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            print(f"  Final batch flush error: {e}")
            stats['errors'] += len(batch)
            stats['imported'] -= len(batch)

    if not dry_run:
        db.session.commit()

    wb.close()

    print(f"\n  Results:")
    print(f"    Imported:     {stats['imported']}")
    print(f"    Skipped:      {stats['skipped']} (already exist)")
    print(f"    Errors:       {stats['errors']}")
    print(f"    EKKA matches: {stats['ekka_matched']} / 148 MKO entries")


# --- CLI ---

def parse_args():
    parser = argparse.ArgumentParser(description='Import Attica structures from Excel + EKKA data')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing to DB')
    parser.add_argument('--clear-existing', action='store_true', help='Delete previously imported Attica structures first')
    parser.add_argument('--excel', default=None, help='Path to Excel file (default: auto-detect in project root)')
    parser.add_argument('--ekka', default=None, help='Path to EKKA CSV file (default: auto-detect in project root)')
    return parser.parse_args()


def main():
    args = parse_args()
    app = create_app()

    with app.app_context():
        print("Import Attica Structures")
        print("=" * 50)
        if args.dry_run:
            print("DRY RUN — no changes will be made\n")

        # Resolve file paths
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
        excel_path = args.excel or os.path.join(project_root, 'mapAttica-Αποτελέσματα Αναζήτησης.xlsx')
        ekka_path = args.ekka or os.path.join(project_root, 'Πίνακας.csv')

        # Step 1: Ensure structure types
        print("\n[1/3] Structure Types")
        types = ensure_types(dry_run=args.dry_run)

        # Step 2: Load EKKA data
        print("\n[2/3] EKKA Data")
        ekka = load_ekka_data(ekka_path)

        # Step 3: Import structures
        print("\n[3/3] Importing Structures")
        import_structures(excel_path, types, ekka,
                          dry_run=args.dry_run,
                          clear_existing=args.clear_existing)

    print("\nDone.")


if __name__ == '__main__':
    main()
