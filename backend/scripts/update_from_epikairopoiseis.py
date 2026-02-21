#!/usr/bin/env python
"""
Update structures from ΕΠΙΚΑΙΡΟΠΟΙΗΣΕΙΣ ΑΡΧΕΙΩΝ ΠΑΡΑΤΗΡΗΤΗΡΙΟΥ.xlsx

Reads 10 sheets with updated contact info, addresses, capacity data.
Creates KAA + MPP structure types and imports 73 new structures.
Fuzzy-matches Excel names to DB names to update existing structures.

Usage:
    python scripts/update_from_epikairopoiseis.py              # Dry run (default)
    python scripts/update_from_epikairopoiseis.py --apply      # Write to DB
"""
import argparse
import os
import re
import sys
from collections import Counter

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import load_workbook

from my_project import create_app
from my_project.extensions import db
from my_project.registry.models import Structure, StructureType


# --- Constants ---

# New structure types to create
NEW_TYPES = {
    'KAA': ('Κέντρο Αποκατάστασης-Αποθεραπείας',
            'Κέντρα Αποκατάστασης-Αποθεραπείας (ΚΑΑ) και ΚΑΑ ΑΜΕΑ'),
    'MPP': ('Μονάδα Παιδικής Προστασίας',
            'Μονάδες Παιδικής Προστασίας (ΜΠΠ)'),
}

# PE mapping: Greek name → 2-letter code
PE_MAP = {
    'ΚΕΝΤΡΙΚΟΣ ΤΟΜΕΑΣ': 'KT',
    'ΒΟΡΕΙΟΣ ΤΟΜΕΑΣ': 'BT',
    'ΝΟΤΙΟΣ ΤΟΜΕΑΣ': 'NT',
    'ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ': 'AA',
    'ΠΕΙΡΑΙΩΣ ΚΑΙ ΝΗΣΩΝ': 'PN',
    'ΔΥΤΙΚΟΣ ΤΟΜΕΑΣ': 'DT',
    'ΔΥΤΙΚΗ ΑΤΤΙΚΗ': 'DA',
}

# Sheet configurations: sheet_name → column mapping (0-indexed from headers row)
# pe=1, city=2 are always the same; data starts at min_row=4
SHEET_CONFIG = {
    'ΚΑΑ ΚΑΑ-ΜΕΑ': {
        'type_code': 'KAA',
        'name_col': 6, 'foreas_col': 4, 'ownership_col': 5,
        'license_col': 7, 'capacity_col': 8, 'address_col': 10,
        'phone_col': 11, 'contact_col': 12, 'email_col': 13,
    },
    'Δημοτικοί Παιδικοί Σταθμοί': {
        'type_code': 'DS',
        'name_col': 4, 'foreas_col': None, 'ownership_col': None,
        'license_col': None, 'capacity_col': 10, 'address_col': 5,
        'phone_col': 15, 'contact_col': None, 'email_col': None,
    },
    'ΚΗΦΗ': {
        'type_code': 'KHFH',
        'name_col': 5, 'foreas_col': 3, 'ownership_col': 4,
        'license_col': 6, 'capacity_col': 7, 'address_col': 10,
        'phone_col': 11, 'contact_col': None, 'email_col': 12,
    },
    'ΜΠΠ': {
        'type_code': 'MPP',
        'name_col': 4, 'foreas_col': 3, 'ownership_col': None,
        'license_col': 5, 'capacity_col': 6, 'address_col': 9,
        'phone_col': 10, 'contact_col': None, 'email_col': 11,
    },
    'ΜΦΗ': {
        'type_code': 'MFH',
        'name_col': 4, 'foreas_col': 5, 'ownership_col': 3,
        'license_col': 6, 'capacity_col': 7, 'address_col': 10,
        'phone_col': 11, 'contact_col': None, 'email_col': 12,
    },
    'ΚΔΑΠ': {
        'type_code': 'KDAP',
        'name_col': 4, 'foreas_col': 5, 'ownership_col': 3,
        'license_col': 6, 'capacity_col': 7, 'address_col': 10,
        'phone_col': 11, 'contact_col': 12, 'email_col': 13,
    },
    'ΚΔΑΠ-ΜΕΑ': {
        'type_code': 'KDAP-AMEA',
        'name_col': 4, 'foreas_col': 5, 'ownership_col': 3,
        'license_col': 6, 'capacity_col': 7, 'address_col': 11,
        'phone_col': 12, 'contact_col': 13, 'email_col': 14,
    },
    'ΣΥΔ': {
        'type_code': 'SYD',
        'name_col': 6, 'foreas_col': 4, 'ownership_col': 5,
        'license_col': 7, 'capacity_col': 8, 'address_col': 11,
        'phone_col': 12, 'contact_col': 13, 'email_col': 14,
    },
    'ΠΙΣΤΟΠΟΙΗΜΕΝΟΙ ΦΟΡΕΙΣ': {
        'type_code': 'MKO',
        'name_col': 3, 'foreas_col': None, 'ownership_col': 5,
        'license_col': 4, 'capacity_col': None, 'address_col': 13,
        'phone_col': 14, 'contact_col': 15, 'email_col': 16,
    },
    'ΜΦΠΑΔ': {
        'type_code': 'MFPAD',
        'name_col': 4, 'foreas_col': 5, 'ownership_col': 3,
        'license_col': 7, 'capacity_col': 10, 'address_col': 16,
        'phone_col': 17, 'contact_col': None, 'email_col': 18,
    },
}


# --- Helpers ---

def normalize_name(s):
    """Aggressive normalization for matching."""
    if not s:
        return ''
    s = str(s).strip().upper()
    s = s.replace('\u201c', '').replace('\u201d', '').replace('"', '')
    s = s.replace('\u00ab', '').replace('\u00bb', '')
    s = s.replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s)
    s = s.strip(' .')
    return s


def parse_address(raw_address):
    """Parse combined Greek address into (street, postal_code) tuple."""
    if not raw_address:
        return None, None

    raw_address = str(raw_address).strip()
    postal_match = re.search(r'\b(\d{3})\s?(\d{2})\b', raw_address)
    postal_code = None
    if postal_match:
        postal_code = postal_match.group(1) + postal_match.group(2)

    parts = [p.strip() for p in raw_address.split(',')]
    street = parts[0] if parts else raw_address

    return street[:200] if street else None, postal_code


def infer_ownership(type_code, raw_ownership, structure_name):
    """Infer ownership from type, explicit column, or name."""
    # Check explicit ownership column first
    if raw_ownership:
        own_str = str(raw_ownership).strip().upper()
        if 'ΔΗΜΟΣ' in own_str or 'ΔΗΜΟΤΙΚ' in own_str or 'ΔΗΜΟΥ' in own_str or 'NPDD' in own_str or 'ΝΠΔΔ' in own_str:
            return 'municipal'
        if 'ΙΔΙΩΤΙΚ' in own_str or 'NPID' in own_str or 'ΝΠΙΔ' in own_str:
            return 'private'
        if 'ΜΚΟ' in own_str or 'NGO' in own_str:
            return 'ngo'

    # Type-based defaults
    if type_code == 'DS':
        return 'municipal'
    if type_code == 'MKO':
        return 'ngo'

    # Name-based fallback
    name_upper = (structure_name or '').upper()
    if 'ΔΗΜΟΣ' in name_upper or 'ΔΗΜΟΤΙΚ' in name_upper or 'ΔΗΜΟΥ' in name_upper:
        return 'municipal'

    return 'private'


def parse_phone(raw_phone):
    """Extract first phone number token, max 20 chars."""
    if not raw_phone:
        return None
    raw = str(int(raw_phone)) if isinstance(raw_phone, float) else str(raw_phone)
    raw = raw.strip()
    if not raw or raw.lower() == 'none':
        return None
    return raw.split()[0][:20]


def parse_capacity(raw_cap):
    """Extract leading integer from capacity field."""
    if not raw_cap:
        return None
    cap_match = re.match(r'(\d+)', str(raw_cap).strip())
    return int(cap_match.group(1)) if cap_match else None


def parse_license(raw_license):
    """Clean and truncate license number."""
    if not raw_license:
        return None
    val = str(raw_license).strip()
    if not val or val in ('', ';;', '-', 'None'):
        return None
    return val[:100]


def parse_email(raw_email):
    """Clean and truncate email."""
    if not raw_email:
        return None
    val = str(raw_email).strip()
    if not val or val.lower() == 'none':
        return None
    return val[:120]


def cell_val(row, col_idx):
    """Safely get cell value from a row tuple by 0-based column index."""
    if col_idx is None or col_idx >= len(row):
        return None
    val = row[col_idx]
    if val is None:
        return None
    return val


def resolve_pe(raw_pe):
    """Resolve PE name to 2-letter code."""
    if not raw_pe:
        return 'XX'
    pe_upper = str(raw_pe).strip().upper()
    for key, code in PE_MAP.items():
        if key in pe_upper or pe_upper in key:
            return code
    return 'XX'


# --- Core Logic ---

def ensure_types(types_dict, dry_run):
    """Create KAA and MPP structure types if they don't exist."""
    created = []
    for code, (name, desc) in NEW_TYPES.items():
        if code not in types_dict:
            if dry_run:
                print(f"  [DRY] Would create type: {code} — {name}")
                fake = StructureType(code=code, name=name, description=desc)
                fake.id = -1
                types_dict[code] = fake
            else:
                st = StructureType(code=code, name=name, description=desc)
                db.session.add(st)
                created.append(code)

    if created and not dry_run:
        db.session.flush()
        # Refresh types_dict with real IDs
        for st in StructureType.query.filter(StructureType.code.in_(created)).all():
            types_dict[st.code] = st
        print(f"  Created types: {', '.join(created)}")
    elif not created:
        print(f"  All types already exist")

    return types_dict


def build_name_index(structures_by_type):
    """Build normalized-name → Structure lookup per type_code."""
    index = {}  # type_code → {normalized_name → Structure}
    for s in structures_by_type:
        norm = normalize_name(s.name)
        type_code = s.structure_type.code if s.structure_type else None
        if type_code not in index:
            index[type_code] = {}
        index[type_code][norm] = s
    return index


def find_match(excel_name, type_index):
    """Try to find matching DB structure by name.

    Returns (structure, match_type) or (None, None).
    match_type is 'exact' or 'fuzzy'.
    """
    norm = normalize_name(excel_name)
    if not norm:
        return None, None

    # Exact match
    if norm in type_index:
        return type_index[norm], 'exact'

    # Substring match — either direction
    for db_norm, structure in type_index.items():
        if len(norm) > 5 and len(db_norm) > 5:
            if norm in db_norm or db_norm in norm:
                return structure, 'fuzzy'

    return None, None


def update_structure(structure, fields, dry_run):
    """Update non-empty fields on a matched structure. Returns True if anything changed."""
    changed = False
    for attr, new_val in fields.items():
        if new_val is None:
            continue
        if isinstance(new_val, str) and not new_val.strip():
            continue
        old_val = getattr(structure, attr, None)
        if old_val != new_val:
            if not dry_run:
                setattr(structure, attr, new_val)
            changed = True
    return changed


def process_sheet(ws, config, types_dict, name_index, seq_counters,
                  existing_codes, dry_run):
    """Process one Excel sheet. Returns stats dict."""
    type_code = config['type_code']
    type_obj = types_dict.get(type_code)
    if not type_obj:
        print(f"    ERROR: Type {type_code} not found in DB")
        return {'exact': 0, 'fuzzy': 0, 'updated': 0, 'inserted': 0, 'unchanged': 0, 'errors': 0}

    type_index = name_index.get(type_code, {})
    stats = {'exact': 0, 'fuzzy': 0, 'updated': 0, 'inserted': 0,
             'unchanged': 0, 'errors': 0}
    batch = []

    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        row_num = i + 4
        try:
            # Read name — required field
            name = cell_val(row, config['name_col'])
            if not name or not str(name).strip():
                continue

            name = str(name).strip()[:300]

            # Read PE and city (always cols 1, 2)
            raw_pe = cell_val(row, 1)
            city = cell_val(row, 2)
            pe_code = resolve_pe(raw_pe)

            # Parse all optional fields
            raw_foreas = cell_val(row, config['foreas_col'])
            raw_ownership = cell_val(row, config['ownership_col'])
            raw_license = cell_val(row, config['license_col'])
            raw_capacity = cell_val(row, config['capacity_col'])
            raw_address = cell_val(row, config['address_col'])
            raw_phone = cell_val(row, config['phone_col'])
            raw_contact = cell_val(row, config['contact_col'])
            raw_email = cell_val(row, config['email_col'])

            # Parse fields
            street, postal_code = parse_address(raw_address)
            phone = parse_phone(raw_phone)
            email = parse_email(raw_email)
            capacity = parse_capacity(raw_capacity)
            license_number = parse_license(raw_license)
            ownership = infer_ownership(type_code, raw_ownership, name)
            city_str = str(city).strip()[:100] if city else None

            # Representative name: prefer contact_col, fall back to foreas_col
            rep_name = None
            if raw_contact:
                rep_name = str(raw_contact).strip()[:200]
            elif raw_foreas:
                rep_name = str(raw_foreas).strip()[:200]

            # Try to match existing structure
            match, match_type = find_match(name, type_index)

            if match:
                # Update existing structure
                if match_type == 'exact':
                    stats['exact'] += 1
                else:
                    stats['fuzzy'] += 1

                fields = {
                    'street': street,
                    'postal_code': postal_code,
                    'representative_phone': phone,
                    'representative_email': email,
                    'representative_name': rep_name,
                    'capacity': capacity,
                    'license_number': license_number,
                    'ownership': ownership,
                }
                # Only update city if we have it
                if city_str:
                    fields['city'] = city_str
                # Update PE if available
                if raw_pe and pe_code != 'XX':
                    pe_full = str(raw_pe).strip()[:100]
                    fields['peripheral_unit'] = pe_full

                if update_structure(match, fields, dry_run):
                    stats['updated'] += 1
                else:
                    stats['unchanged'] += 1
            else:
                # Insert new structure
                seq_key = f"{type_code}-{pe_code}"
                seq_counters[seq_key] += 1
                code = f"{type_code}-{pe_code}-{seq_counters[seq_key]:04d}"

                # Avoid code collision
                while code in existing_codes:
                    seq_counters[seq_key] += 1
                    code = f"{type_code}-{pe_code}-{seq_counters[seq_key]:04d}"

                if dry_run:
                    if stats['inserted'] < 5:
                        print(f"    [DRY] NEW: {code} | {name[:60]}")
                    elif stats['inserted'] == 5:
                        print(f"    ... (more new structures)")
                    stats['inserted'] += 1
                    existing_codes.add(code)
                    # Add to index so duplicates within sheet don't re-insert
                    type_index[normalize_name(name)] = None
                    continue

                structure = Structure(
                    code=code,
                    type_id=type_obj.id,
                    name=name,
                    street=street,
                    city=city_str,
                    postal_code=postal_code,
                    representative_name=rep_name,
                    representative_phone=phone,
                    representative_email=email,
                    capacity=capacity,
                    status='active',
                    ownership=ownership,
                    license_number=license_number,
                    peripheral_unit=str(raw_pe).strip()[:100] if raw_pe else None,
                )
                db.session.add(structure)
                existing_codes.add(code)
                batch.append(structure)
                stats['inserted'] += 1

                # Add to index so we don't try to re-insert within same sheet
                type_index[normalize_name(name)] = structure

                # Batch flush
                if len(batch) >= 200:
                    try:
                        db.session.flush()
                    except Exception as e:
                        db.session.rollback()
                        print(f"    Batch flush error at row {row_num}: {e}")
                        stats['errors'] += len(batch)
                        stats['inserted'] -= len(batch)
                    batch.clear()

        except Exception as e:
            print(f"    Row {row_num}: ERROR — {e}")
            stats['errors'] += 1

    # Final flush for this sheet
    if batch and not dry_run:
        try:
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            print(f"    Final flush error: {e}")
            stats['errors'] += len(batch)
            stats['inserted'] -= len(batch)

    return stats


def init_seq_counters():
    """Scan existing structure codes to find max sequence per type-PE combo."""
    counters = Counter()
    for (code,) in db.session.query(Structure.code).all():
        parts = code.rsplit('-', 1)
        if len(parts) == 2:
            prefix = parts[0]  # e.g. "KAA-KT" or "KDAP-AMEA-BT"
            try:
                seq = int(parts[1])
                if seq > counters[prefix]:
                    counters[prefix] = seq
            except ValueError:
                pass
    return counters


# --- Main ---

def main():
    parser = argparse.ArgumentParser(
        description='Update structures from ΕΠΙΚΑΙΡΟΠΟΙΗΣΕΙΣ ΑΡΧΕΙΩΝ ΠΑΡΑΤΗΡΗΤΗΡΙΟΥ')
    parser.add_argument('--apply', action='store_true',
                        help='Actually write to DB (default is dry run)')
    parser.add_argument('--excel', default=None,
                        help='Path to Excel file (default: auto-detect)')
    args = parser.parse_args()
    dry_run = not args.apply

    app = create_app()

    with app.app_context():
        print("=" * 60)
        print("Update Structures from ΕΠΙΚΑΙΡΟΠΟΙΗΣΕΙΣ ΑΡΧΕΙΩΝ ΠΑΡΑΤΗΡΗΤΗΡΙΟΥ")
        print("=" * 60)
        if dry_run:
            print("DRY RUN — pass --apply to write changes\n")
        else:
            print("APPLYING CHANGES\n")

        # Resolve Excel path
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
        excel_path = args.excel or os.path.join(
            project_root, 'ΕΠΙΚΑΙΡΟΠΟΙΗΣΕΙΣ ΑΡΧΕΙΩΝ ΠΑΡΑΤΗΡΗΤΗΡΙΟΥ.xlsx')
        if not os.path.exists(excel_path):
            print(f"ERROR: Excel file not found: {excel_path}")
            sys.exit(1)
        print(f"Excel: {os.path.basename(excel_path)}\n")

        # Step 1: Ensure structure types
        print("[1/4] Structure Types")
        types_dict = {st.code: st for st in StructureType.query.all()}
        print(f"  Existing: {len(types_dict)} types ({', '.join(sorted(types_dict.keys()))})")
        types_dict = ensure_types(types_dict, dry_run)

        # Step 2: Build name index from DB
        print("\n[2/4] Loading existing structures")
        all_structures = Structure.query.options(
            db.joinedload(Structure.structure_type)
        ).all()
        print(f"  {len(all_structures)} structures in DB")
        name_index = build_name_index(all_structures)

        # Step 3: Init sequence counters from existing codes
        seq_counters = init_seq_counters()
        existing_codes = set(s.code for s in all_structures)
        print(f"  {len(existing_codes)} existing codes tracked")

        # Step 4: Process each sheet
        print("\n[3/4] Processing sheets")
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        available_sheets = wb.sheetnames
        print(f"  Sheets in workbook: {available_sheets}\n")

        totals = {'exact': 0, 'fuzzy': 0, 'updated': 0, 'inserted': 0,
                  'unchanged': 0, 'errors': 0}

        for sheet_name, config in SHEET_CONFIG.items():
            # Find the sheet — prefer exact match, then contained match
            ws = None
            # Pass 1: exact match
            for available in available_sheets:
                if sheet_name == available or sheet_name == available.strip():
                    ws = wb[available]
                    break
            # Pass 2: partial match — pick the best (longest match)
            if ws is None:
                best_match = None
                best_len = 0
                for available in available_sheets:
                    avail_stripped = available.strip()
                    if sheet_name in avail_stripped or avail_stripped in sheet_name:
                        match_len = min(len(sheet_name), len(avail_stripped))
                        if match_len > best_len:
                            best_len = match_len
                            best_match = available
                if best_match:
                    ws = wb[best_match]

            if ws is None:
                print(f"  --- {config['type_code']} ({sheet_name}) --- SHEET NOT FOUND, skipping")
                continue

            print(f"  === {config['type_code']} ({ws.title}) ===")
            stats = process_sheet(ws, config, types_dict, name_index,
                                  seq_counters, existing_codes, dry_run)

            matched = stats['exact'] + stats['fuzzy']
            print(f"    Matched: {matched} (exact={stats['exact']}, fuzzy={stats['fuzzy']})")
            print(f"    Updated: {stats['updated']}, Unchanged: {stats['unchanged']}")
            print(f"    New:     {stats['inserted']}")
            if stats['errors']:
                print(f"    Errors:  {stats['errors']}")
            print()

            for k in totals:
                totals[k] += stats[k]

        wb.close()

        # Commit
        if not dry_run:
            db.session.commit()
            print("[4/4] Changes committed to DB")
        else:
            print("[4/4] Dry run complete — no changes made")

        # Summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        total_matched = totals['exact'] + totals['fuzzy']
        print(f"  Matched:   {total_matched} (exact={totals['exact']}, fuzzy={totals['fuzzy']})")
        print(f"  Updated:   {totals['updated']}")
        print(f"  Unchanged: {totals['unchanged']}")
        print(f"  Inserted:  {totals['inserted']}")
        if totals['errors']:
            print(f"  Errors:    {totals['errors']}")
        print()


if __name__ == '__main__':
    main()
