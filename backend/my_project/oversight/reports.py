"""Report generation utilities (PDF/XLSX).

Each generate_* function returns bytes (the file content).
The route handler wraps these in a Flask Response with the correct mimetype.
"""
import io
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..extensions import db
from ..registry.models import Structure, StructureType, License, Sanction
from ..inspections.models import Inspection, InspectionReport
from ..sanctions.models import SanctionDecision
from .models import SocialAdvisorReport


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

HEADER_BG = colors.HexColor('#1a3aa3')
HEADER_FG = colors.white
ALT_ROW = colors.HexColor('#f5f2ec')

XL_HEADER_FILL = PatternFill('solid', fgColor='1a3aa3')
XL_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
XL_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'],
        fontSize=16, spaceAfter=12, textColor=colors.HexColor('#2a2520'),
    ))
    styles.add(ParagraphStyle(
        'ReportSubtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#6b6560'), spaceAfter=16,
    ))
    return styles


def _pdf_table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), HEADER_FG),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e8e2d8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ALT_ROW]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ])


def _xl_write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = XL_HEADER_FONT
        cell.fill = XL_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = XL_THIN_BORDER


def _xl_auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _fmt_date(d):
    if not d:
        return '—'
    if isinstance(d, str):
        d = date.fromisoformat(d)
    return d.strftime('%d/%m/%Y')


# ---------------------------------------------------------------------------
# Report: Registry Status
# ---------------------------------------------------------------------------

def _registry_data():
    structures = Structure.query.order_by(Structure.name).all()
    headers = ['Κωδικός', 'Επωνυμία', 'Τύπος', 'Κατάσταση', 'Πόλη', 'Λήξη Αδείας']
    rows = []
    for s in structures:
        rows.append([
            s.code,
            s.name,
            s.type.name if s.type else '—',
            s.status,
            s.city or '—',
            _fmt_date(s.license_expiry),
        ])
    return headers, rows


def generate_registry_pdf():
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = _pdf_styles()
    elements = []

    elements.append(Paragraph('Μητρώο Δομών — Κατάσταση', styles['ReportTitle']))
    elements.append(Paragraph(f'Ημερομηνία: {_fmt_date(date.today())}', styles['ReportSubtitle']))

    headers, rows = _registry_data()
    data = [headers] + rows
    col_widths = [55, 130, 80, 60, 60, 65]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()


def generate_registry_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Μητρώο Δομών'

    headers, rows = _registry_data()
    _xl_write_header(ws, headers)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = XL_THIN_BORDER
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Report: Inspections
# ---------------------------------------------------------------------------

def _inspections_data(date_from=None, date_to=None):
    from sqlalchemy.orm import joinedload
    query = Inspection.query.options(joinedload(Inspection.structure)).order_by(Inspection.scheduled_date.desc())
    if date_from:
        query = query.filter(Inspection.scheduled_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(Inspection.scheduled_date <= date.fromisoformat(date_to))
    inspections = query.all()

    headers = ['Ημ/νία', 'Τύπος', 'Κατάσταση', 'Συμπέρασμα', 'Δομή']
    rows = []
    for insp in inspections:
        structure = insp.structure
        rows.append([
            _fmt_date(insp.scheduled_date),
            insp.type,
            insp.status,
            insp.conclusion or '—',
            structure.name if structure else '—',
        ])
    return headers, rows


def generate_inspections_pdf(date_from=None, date_to=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = _pdf_styles()
    elements = []

    elements.append(Paragraph('Αναφορά Ελέγχων', styles['ReportTitle']))
    subtitle = f'Ημερομηνία: {_fmt_date(date.today())}'
    if date_from or date_to:
        subtitle += f'  |  Περίοδος: {date_from or "—"} — {date_to or "—"}'
    elements.append(Paragraph(subtitle, styles['ReportSubtitle']))

    headers, rows = _inspections_data(date_from, date_to)
    data = [headers] + rows
    col_widths = [65, 70, 70, 90, 130]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()


def generate_inspections_xlsx(date_from=None, date_to=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Έλεγχοι'

    headers, rows = _inspections_data(date_from, date_to)
    _xl_write_header(ws, headers)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = XL_THIN_BORDER
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Report: Sanctions
# ---------------------------------------------------------------------------

def _sanctions_data(date_from=None, date_to=None):
    from sqlalchemy.orm import joinedload
    query = Sanction.query.options(joinedload(Sanction.structure)).order_by(Sanction.imposed_date.desc())
    if date_from:
        query = query.filter(Sanction.imposed_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(Sanction.imposed_date <= date.fromisoformat(date_to))
    sanctions = query.all()

    headers = ['Ημ/νία', 'Τύπος', 'Ποσό', 'Κατάσταση', 'Δομή']
    rows = []
    for s in sanctions:
        structure = s.structure
        rows.append([
            _fmt_date(s.imposed_date),
            s.type,
            f'{s.amount} €' if s.amount else '—',
            s.status,
            structure.name if structure else '—',
        ])
    return headers, rows


def generate_sanctions_pdf(date_from=None, date_to=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = _pdf_styles()
    elements = []

    elements.append(Paragraph('Αναφορά Κυρώσεων', styles['ReportTitle']))
    subtitle = f'Ημερομηνία: {_fmt_date(date.today())}'
    if date_from or date_to:
        subtitle += f'  |  Περίοδος: {date_from or "—"} — {date_to or "—"}'
    elements.append(Paragraph(subtitle, styles['ReportSubtitle']))

    headers, rows = _sanctions_data(date_from, date_to)
    data = [headers] + rows
    col_widths = [65, 80, 60, 70, 130]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()


def generate_sanctions_xlsx(date_from=None, date_to=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Κυρώσεις'

    headers, rows = _sanctions_data(date_from, date_to)
    _xl_write_header(ws, headers)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = XL_THIN_BORDER
    _xl_auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Report: Sanction Decisions (workflow)
# ---------------------------------------------------------------------------

DECISION_STATUS_LABELS = {
    'draft': 'Πρόχειρη',
    'submitted': 'Υποβληθείσα',
    'approved': 'Εγκεκριμένη',
    'returned': 'Επιστράφηκε',
    'notified': 'Κοινοποιήθηκε',
    'paid': 'Πληρώθηκε',
    'appealed': 'Ένσταση',
    'overdue': 'Εκπρόθεσμη',
    'cancelled': 'Ακυρώθηκε',
}


def _decisions_data(date_from=None, date_to=None):
    from sqlalchemy.orm import joinedload
    query = SanctionDecision.query.options(
        joinedload(SanctionDecision.sanction).joinedload(Sanction.structure)
    ).order_by(SanctionDecision.created_at.desc())
    if date_from:
        query = query.filter(SanctionDecision.created_at >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(SanctionDecision.created_at <= date.fromisoformat(date_to))
    decisions = query.all()

    headers = ['Α/Α', 'Πρωτόκολλο', 'Δομή', 'Παράβαση', 'Ποσό (€)', 'Κατάσταση', 'Λήξη Πληρωμής']
    rows = []
    for d in decisions:
        sanction = d.sanction
        structure = sanction.structure if sanction else None
        rows.append([
            d.id,
            d.protocol_number or '—',
            structure.name if structure else '—',
            d.violation_code or '—',
            f'{d.final_amount:,.2f}' if d.final_amount else '—',
            DECISION_STATUS_LABELS.get(d.status, d.status),
            _fmt_date(d.payment_deadline),
        ])
    return headers, rows, decisions


def generate_decisions_pdf(date_from=None, date_to=None):
    from sqlalchemy import func

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = _pdf_styles()
    elements = []

    elements.append(Paragraph('Αναφορά Αποφάσεων Κυρώσεων', styles['ReportTitle']))
    subtitle = f'Ημερομηνία: {_fmt_date(date.today())}'
    if date_from or date_to:
        subtitle += f'  |  Περίοδος: {date_from or "—"} — {date_to or "—"}'
    elements.append(Paragraph(subtitle, styles['ReportSubtitle']))

    headers, rows, decisions = _decisions_data(date_from, date_to)

    # Summary stats
    total_amount = sum(d.final_amount or 0 for d in decisions)
    paid_amount = sum(d.paid_amount or 0 for d in decisions if d.status == 'paid')
    summary_text = (
        f'Σύνολο αποφάσεων: {len(decisions)} | '
        f'Συνολικό ποσό: {total_amount:,.2f}€ | '
        f'Εισπράξεις: {paid_amount:,.2f}€'
    )
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 10))

    data = [headers] + rows
    col_widths = [30, 65, 110, 70, 55, 65, 60]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()


def generate_decisions_xlsx(date_from=None, date_to=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Αποφάσεις Κυρώσεων'

    headers, rows, decisions = _decisions_data(date_from, date_to)
    _xl_write_header(ws, headers)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = XL_THIN_BORDER
    _xl_auto_width(ws)

    # Summary row
    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Σύνολο:').font = Font(bold=True)
    total_amount = sum(d.final_amount or 0 for d in decisions)
    paid_amount = sum(d.paid_amount or 0 for d in decisions if d.status == 'paid')
    ws.cell(row=summary_row, column=5, value=f'{total_amount:,.2f}').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value='Εισπράξεις:').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=5, value=f'{paid_amount:,.2f}').font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

REPORT_GENERATORS = {
    'registry': {
        'pdf': lambda **kw: generate_registry_pdf(),
        'xlsx': lambda **kw: generate_registry_xlsx(),
    },
    'inspections': {
        'pdf': generate_inspections_pdf,
        'xlsx': generate_inspections_xlsx,
    },
    'sanctions': {
        'pdf': generate_sanctions_pdf,
        'xlsx': generate_sanctions_xlsx,
    },
    'decisions': {
        'pdf': generate_decisions_pdf,
        'xlsx': generate_decisions_xlsx,
    },
}
